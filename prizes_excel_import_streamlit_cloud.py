#======SUBJECT TO CHANGE===========
url = "https://fsm.sg.formulasquare.com/fsm_api/wawaji_cms/"
#=================================

import time
import openpyxl
import streamlit as st
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

options = Options()
options.headless = True

st.title("🎁 5G Claw Machine CMS - input reward codes from excel file")
st.write("Please use the form below to enter the codes into the CMS system.")
st.write("If at any point you encounter an error message, this could be due to the following: 1) username and/or password is incorrrect; or 2) reward code already exists in the CMS system.")

uploaded_file = st.file_uploader("Upload file", type=['xlsx'])

form = st.form("reward_codes_excel_form")

username = form.text_input("Please enter the CMS username")
password = form.text_input("Please enter the CMS password")
row_start = form.text_input("Row number that corresponds to where the codes begin (e.g. if a header row exists, please key in 2)")
reward_code_col_num = form.text_input("Column number containing codes (e.g. key in 1 if reward codes are in column A, 2 if in column B, etc.)")
prize_name_col_num = form.text_input("Column number containing the names of the prizes (e.g. key in 2 if prize names are in column B, 3 if in column C, etc.)")
submit = form.form_submit_button("Add prizes to CMS")

if uploaded_file is not None:
    if submit:
        row_start = int(row_start)
        reward_code_col_num = int(reward_code_col_num)
        prize_name_col_num = int(prize_name_col_num)
        st.info("Running. Please do NOT click on the button again.")

        #open workbook
        wrkbk = openpyxl.load_workbook(uploaded_file, data_only=True)
        sh = wrkbk.active

        web=webdriver.Chrome(options=options)
        web.get(url)
        time.sleep(3)

        #Login to CMS
        username_field = web.find_element("xpath", '/html/body/div/div/main/div/div/div/div/form/div[1]/input')
        username_field.send_keys(username)
        password_field = web.find_element("xpath", '/html/body/div/div/main/div/div/div/div/form/div[3]/input')
        password_field.send_keys(password)
        submit_login_button = web.find_element("xpath", '/html/body/div/div/main/div/div/div/div/form/div[5]/input')
        submit_login_button.click()
        time.sleep(2)

        #navigate to reward codes page
        reward_codes = web.find_element("xpath", '/html/body/div/div/main/div/div[1]/button[4]')
        reward_codes.click()
        time.sleep(1)

        for i in range (row_start, sh.max_row+1):
           try:
               #click on add new button
               add_new = web.find_element("xpath", '/html/body/div/div/main/div/div[2]/div/div[1]/button')
               add_new.click()
               time.sleep(1)

               #input reward code
               reward_code_cell_obj = sh.cell(row=i, column=reward_code_col_num)
               reward_code_field = web.find_element("xpath", '/html/body/div/div/main/div/div[2]/div/div/form/div[1]/input')
               reward_code_field.send_keys(reward_code_cell_obj.value)

               #input corresponding prize name
               prize_name_cell_obj = sh.cell(row=i, column=prize_name_col_num)
               prize_name_field = web.find_element("xpath", '/html/body/div/div/main/div/div[2]/div/div/form/div[3]/input')
               prize_name_field.send_keys(prize_name_cell_obj.value)

               #click on submit button to add new prize
               submit_button = web.find_element("xpath", '/html/body/div/div/main/div/div[2]/div/div/form/div[8]/input')
               submit_button.click()
               time.sleep(1)
               st.write(f"Prize code {reward_code_cell_obj.value} - {prize_name_cell_obj.value} has been added to the CMS system")
           except:
               TypeError
               pass
               
        st.success("All prize codes have been added to the CMS system")
